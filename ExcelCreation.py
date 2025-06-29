import openpyxl
from openpyxl.styles import Font
from datetime import datetime, timedelta

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Hospital_Records"

# Define headers
headers = [
    "Name",
    "Phone Number",
    "Address",
    "Age",
    "Gender",
]

# Write headers to the sheet
for col_num, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.font = Font(bold=True)

##CHNAGE PHONE NUMBER HERE and run
sample_data = [
    ["Shravani Shelke", "+919049865451", "24 MG Road, Bengaluru", 28, "Male"],
    ["Vanshika panjwani", "+919049865451", "24 MG Road, Bengaluru", 28, "Male"],
    ["Epshita Ninawe", "+919049865451", "24 MG Road, Bengaluru", 28, "Male"],
     ["Anushka Desai", "+919049865451", "24 MG Road, Bengaluru", 28, "Male"],
]

# Insert each row into the worksheet
for row_num, row_data in enumerate(sample_data, start=2):
    for col_num, value in enumerate(row_data, start=1):
        ws.cell(row=row_num, column=col_num, value=value)

# Auto-adjust column widths
for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column_letter].width = max_length + 2

# Save the workbook
wb.save("Hospital_Records.xlsx")
print(f"Excel sheet 'Hospital_Records.xlsx' created successfully with {row_num} sample rows.")

## Un-comment and run below code for "Appointment_Details.xlsx"

# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "Appointment_Details"
#
# # Define headers
# headers = [
#     "Name",
#     "Appointment Date",
#     "Time Slot",
#     "Age",
#     "Gender",
#     "Phone Number",
#     "Address",
# ]
#
# # Write headers to the sheet
# for col_num, header in enumerate(headers, start=1):
#     cell = ws.cell(row=1, column=col_num, value=header)
#     cell.font = Font(bold=True)
#
# # Auto-adjust column widths
# for column in ws.columns:
#     max_length = 0
#     column_letter = column[0].column_letter
#     for cell in column:
#         if cell.value:
#             max_length = max(max_length, len(str(cell.value)))
#     ws.column_dimensions[column_letter].width = max_length + 2
#
# # Save the workbook
# wb.save("Appointment_Details.xlsx")
# print("Excel sheet 'Hospital_Records.xlsx' created successfully with 5 sample rows.")